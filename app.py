import os
import sqlite3
import secrets
import time
from io import BytesIO
from datetime import date, datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, g, abort
)
from werkzeug.security import generate_password_hash, check_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "college.db")

def get_secret_key():
    env_key = os.environ.get("SECRET_KEY")
    if env_key and env_key not in ("change-this-secret-key-in-production", "ssaq-secret-key-production-change-this"):
        return env_key
    key_path = os.path.join(BASE_DIR, ".secret_key")
    if os.path.exists(key_path):
        try:
            with open(key_path, "r", encoding="utf-8") as f:
                key = f.read().strip()
                if key:
                    return key
        except Exception:
            pass
    new_key = secrets.token_hex(32)
    try:
        with open(key_path, "w", encoding="utf-8") as f:
            f.write(new_key)
    except Exception:
        pass
    return new_key

app = Flask(__name__)
app.config.update(
    SECRET_KEY=get_secret_key(),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=bool(os.environ.get("RENDER") or os.environ.get("HTTPS")),
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    MAX_CONTENT_LENGTH=16 * 1024 * 1024,  # Max 16MB file upload limit
)

# ----------------------------------------------------------------------
# Rate Limiting & Anti-Brute-Force
# ----------------------------------------------------------------------
LOGIN_ATTEMPTS = {}  # ip -> {"count": int, "blocked_until": float, "last_attempt": float}

def is_ip_rate_limited(ip):
    now = time.time()
    record = LOGIN_ATTEMPTS.get(ip)
    if not record:
        return False, 0
    if record.get("blocked_until", 0) > now:
        remaining = int(record["blocked_until"] - now)
        return True, remaining
    if now - record.get("last_attempt", 0) > 300:
        LOGIN_ATTEMPTS.pop(ip, None)
        return False, 0
    return False, 0

def record_failed_login(ip):
    now = time.time()
    record = LOGIN_ATTEMPTS.setdefault(ip, {"count": 0, "blocked_until": 0, "last_attempt": now})
    record["count"] += 1
    record["last_attempt"] = now
    if record["count"] >= 5:
        record["blocked_until"] = now + 300  # Lock out for 5 minutes after 5 failed attempts

def record_successful_login(ip):
    LOGIN_ATTEMPTS.pop(ip, None)

def is_safe_redirect(url):
    if not url:
        return False
    return url.startswith("/") and not url.startswith("//") and "\\" not in url

# ----------------------------------------------------------------------
# CSRF Protection & Security Headers
# ----------------------------------------------------------------------
def generate_csrf_token():
    if "_csrf_token" not in session:
        session["_csrf_token"] = secrets.token_hex(32)
    return session["_csrf_token"]

@app.context_processor
def inject_csrf():
    return {"csrf_token": generate_csrf_token}

@app.before_request
def csrf_protect():
    if request.method in ("POST", "PUT", "DELETE", "PATCH"):
        token = request.form.get("csrf_token") or request.headers.get("X-CSRFToken")
        expected_token = session.get("_csrf_token")
        if not expected_token or not token or not secrets.compare_digest(token, expected_token):
            flash("Security check failed (CSRF token invalid or expired). Please try again.", "error")
            return redirect(request.referrer or url_for("dashboard"))

@app.after_request
def set_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    return response

# ----------------------------------------------------------------------
# Academy & Admin / Developer contact shown in the app footer and on every slip
# ----------------------------------------------------------------------
ACADEMY_NAME = "The Smart Skills Academy Qalagay"
COLLEGE_NAME = ACADEMY_NAME
ACADEMY_AFFILIATION = ""
COLLEGE_AFFILIATION = ACADEMY_AFFILIATION
ADMIN_NAME = "Rashid Zada"
ADMIN_CONTACT = "0347-0983567"
ADMIN_CONTACTS = [ADMIN_CONTACT]
COLLEGE_ADMIN_CONTACTS = ADMIN_CONTACTS

DEVELOPER_NAME = "Rashid Zada"
DEVELOPER_TITLE = "Full Stack Developer"
DEVELOPER_WHATSAPP = "923470983567"          # international format, no +
DEVELOPER_WHATSAPP_DISPLAY = "0347-0983567"

# ----------------------------------------------------------------------
# Database helpers
# ----------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
        ensure_payment_schema(g.db)
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    first_run = not os.path.exists(DB_PATH)
    db = sqlite3.connect(DB_PATH)
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS admin (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            subject TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS courses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            duration TEXT,
            fee REAL NOT NULL DEFAULT 0,
            id_card_fee REAL NOT NULL DEFAULT 0,
            dmc_fee REAL NOT NULL DEFAULT 0,
            exam_fee REAL NOT NULL DEFAULT 0,
            fund_fee REAL NOT NULL DEFAULT 0,
            teacher_id INTEGER,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_no TEXT,
            name TEXT NOT NULL,
            father_name TEXT,
            phone TEXT,
            course_id INTEGER,
            teacher_id INTEGER,
            total_fee REAL NOT NULL DEFAULT 0,
            installment_count INTEGER NOT NULL DEFAULT 1,
            admission_date TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (course_id) REFERENCES courses(id) ON DELETE SET NULL,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            installment_no INTEGER NOT NULL,
            tuition_amount REAL NOT NULL DEFAULT 0,
            id_card_fee REAL NOT NULL DEFAULT 0,
            dmc_fee REAL NOT NULL DEFAULT 0,
            exam_fee REAL NOT NULL DEFAULT 0,
            fund_fee REAL NOT NULL DEFAULT 0,
            due_date TEXT,
            paid INTEGER NOT NULL DEFAULT 0,
            paid_date TEXT,
            FOREIGN KEY (student_id) REFERENCES students(id) ON DELETE CASCADE
        );
        """
    )
    db.commit()
    ensure_payment_schema(db)

    if first_run:
        db.execute(
            "INSERT INTO admin (username, password_hash) VALUES (?, ?)",
            ("admin", generate_password_hash("admin123")),
        )
        db.commit()
    db.close()


def ensure_payment_schema(db):
    columns = [row[1] for row in db.execute("PRAGMA table_info(payments)").fetchall()]
    if not columns:
        return
    if "paid_amount" not in columns:
        db.execute("ALTER TABLE payments ADD COLUMN paid_amount REAL NOT NULL DEFAULT 0")
        db.execute(
            """UPDATE payments
               SET paid_amount = tuition_amount + id_card_fee + dmc_fee + exam_fee + fund_fee
               WHERE paid = 1 AND paid_amount = 0"""
        )
        db.commit()


# ----------------------------------------------------------------------
# Auth
# ----------------------------------------------------------------------
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin_id"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("admin_id"):
        return redirect(url_for("dashboard"))

    client_ip = request.headers.get("X-Forwarded-For", request.remote_addr).split(",")[0].strip()
    is_limited, remaining_secs = is_ip_rate_limited(client_ip)
    if is_limited:
        flash(f"Too many failed login attempts. Please wait {remaining_secs} seconds before trying again.", "error")
        return render_template("login.html", dev=dev_context()), 429

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        db = get_db()
        row = db.execute("SELECT * FROM admin WHERE LOWER(username) = LOWER(?)", (username,)).fetchone()
        if not row and username.lower() in ["rashid", "rashid zada", "rashidzada", "admin"]:
            row = db.execute("SELECT * FROM admin LIMIT 1").fetchone()

        if row and check_password_hash(row["password_hash"], password):
            csrf_token = session.get("_csrf_token")
            session.clear()
            if csrf_token:
                session["_csrf_token"] = csrf_token
            session["admin_id"] = row["id"]
            session["admin_username"] = row["username"]
            record_successful_login(client_ip)
            flash("Welcome back!", "success")
            nxt = request.args.get("next")
            if not is_safe_redirect(nxt):
                nxt = url_for("dashboard")
            return redirect(nxt)

        record_failed_login(client_ip)
        time.sleep(0.4)
        flash("Invalid username or password.", "error")

    return render_template("login.html", dev=dev_context())


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


def dev_context():
    return {
        "name": DEVELOPER_NAME,
        "title": DEVELOPER_TITLE,
        "whatsapp": DEVELOPER_WHATSAPP,
        "whatsapp_display": DEVELOPER_WHATSAPP_DISPLAY,
        "academy": ACADEMY_NAME,
        "college": ACADEMY_NAME,
        "affiliation": ACADEMY_AFFILIATION,
        "admin_name": ADMIN_NAME,
        "admin_contact": ADMIN_CONTACT,
        "admin_contacts": ADMIN_CONTACTS,
        "admin_display": f"{ADMIN_NAME} ({ADMIN_CONTACT})",
    }


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def format_pk_whatsapp(phone):
    """Normalize a Pakistani phone number into wa.me format (92XXXXXXXXXX)."""
    if not phone:
        return ""
    digits = "".join(ch for ch in phone if ch.isdigit())
    if digits.startswith("0"):
        digits = "92" + digits[1:]
    elif digits.startswith("92"):
        pass
    elif len(digits) == 10:
        digits = "92" + digits
    return digits


def safe_filename_part(value):
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in str(value or "").strip())
    cleaned = "_".join(part for part in cleaned.split("_") if part)
    return cleaned or "student"


PAYMENT_TOTAL_SQL = "(tuition_amount + id_card_fee + dmc_fee + exam_fee + fund_fee)"
PAYMENT_REMAINING_SQL = f"MAX({PAYMENT_TOTAL_SQL} - COALESCE(paid_amount, 0), 0)"
PAYMENT_TOTAL_P_SQL = "(p.tuition_amount + p.id_card_fee + p.dmc_fee + p.exam_fee + p.fund_fee)"
PAYMENT_REMAINING_P_SQL = f"MAX({PAYMENT_TOTAL_P_SQL} - COALESCE(p.paid_amount, 0), 0)"


def payment_total(payment):
    return (
        float(payment["tuition_amount"] or 0)
        + float(payment["id_card_fee"] or 0)
        + float(payment["dmc_fee"] or 0)
        + float(payment["exam_fee"] or 0)
        + float(payment["fund_fee"] or 0)
    )


def payment_remaining(payment):
    return max(payment_total(payment) - float(payment["paid_amount"] or 0), 0)


def get_student_payment_rows(db, student_id):
    return db.execute(
        f"""SELECT *,
                  {PAYMENT_TOTAL_P_SQL} AS payment_total,
                  {PAYMENT_REMAINING_P_SQL} AS remaining_amount
           FROM payments p
           WHERE student_id = ?
           ORDER BY installment_no""",
        (student_id,),
    ).fetchall()


def summarize_payments(payments):
    total_payable = sum(float(p["payment_total"] or 0) for p in payments)
    total_paid = sum(float(p["paid_amount"] or 0) for p in payments)
    total_due = sum(float(p["remaining_amount"] or 0) for p in payments)
    return total_payable, total_paid, total_due


def create_payments_for_student(db, student_id, total_fee, installment_count, course, admission_date_str):
    installment_amount = round(total_fee / installment_count, 2)
    try:
        base_date = datetime.strptime(admission_date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        base_date = date.today()

    for i in range(1, installment_count + 1):
        due = base_date + timedelta(days=30 * (i - 1)) if i > 1 else base_date + timedelta(days=15)
        id_card = course["id_card_fee"] if course and i == 1 else 0
        dmc = course["dmc_fee"] if course and i == 1 else 0
        exam = course["exam_fee"] if course and i == 1 else 0
        fund = course["fund_fee"] if course and i == 1 else 0
        db.execute(
            """INSERT INTO payments
               (student_id, installment_no, tuition_amount, id_card_fee, dmc_fee, exam_fee, fund_fee, due_date, paid)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0)""",
            (student_id, i, installment_amount, id_card, dmc, exam, fund, due.isoformat()),
        )
    db.commit()


def rebuild_unpaid_payments_for_student(db, student_id, total_fee, installment_count, course, admission_date_str):
    paid_rows = db.execute(
        "SELECT installment_no FROM payments WHERE student_id = ? AND (paid = 1 OR paid_amount > 0)",
        (student_id,),
    ).fetchall()
    paid_installments = {row["installment_no"] for row in paid_rows}
    if paid_installments:
        installment_count = max(installment_count, max(paid_installments))

    db.execute("DELETE FROM payments WHERE student_id = ? AND paid = 0 AND paid_amount = 0", (student_id,))

    installment_amount = round(total_fee / installment_count, 2)
    try:
        base_date = datetime.strptime(admission_date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        base_date = date.today()

    for i in range(1, installment_count + 1):
        if i in paid_installments:
            continue
        due = base_date + timedelta(days=30 * (i - 1)) if i > 1 else base_date + timedelta(days=15)
        id_card = course["id_card_fee"] if course and i == 1 else 0
        dmc = course["dmc_fee"] if course and i == 1 else 0
        exam = course["exam_fee"] if course and i == 1 else 0
        fund = course["fund_fee"] if course and i == 1 else 0
        db.execute(
            """INSERT INTO payments
               (student_id, installment_no, tuition_amount, id_card_fee, dmc_fee, exam_fee, fund_fee, due_date, paid)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0)""",
            (student_id, i, installment_amount, id_card, dmc, exam, fund, due.isoformat()),
        )
    db.commit()


# ----------------------------------------------------------------------
# Dashboard
# ----------------------------------------------------------------------
@app.route("/")
@login_required
def dashboard():
    db = get_db()
    total_students = db.execute("SELECT COUNT(*) c FROM students").fetchone()["c"]
    total_teachers = db.execute("SELECT COUNT(*) c FROM teachers").fetchone()["c"]
    total_courses = db.execute("SELECT COUNT(*) c FROM courses").fetchone()["c"]

    collected = db.execute(
        "SELECT COALESCE(SUM(paid_amount), 0) s FROM payments"
    ).fetchone()["s"]
    pending = db.execute(
        f"SELECT COALESCE(SUM({PAYMENT_REMAINING_SQL}), 0) s FROM payments"
    ).fetchone()["s"]
    total_expected = collected + pending

    dues = db.execute(
        f"""SELECT p.*, s.name AS student_name, s.candidate_no, s.phone,
                  c.name AS course_name,
                  {PAYMENT_TOTAL_P_SQL} AS payment_total,
                  {PAYMENT_REMAINING_P_SQL} AS remaining_amount
           FROM payments p
           JOIN students s ON s.id = p.student_id
           LEFT JOIN courses c ON c.id = s.course_id
           WHERE {PAYMENT_REMAINING_P_SQL} > 0
           ORDER BY p.due_date ASC
           LIMIT 25"""
    ).fetchall()

    dues_total_paid = sum(float(d["paid_amount"] or 0) for d in dues)
    dues_total_remaining = sum(float(d["remaining_amount"] or 0) for d in dues)

    courses = db.execute(
        f"""SELECT c.*, t.name AS teacher_name,
                  (SELECT COUNT(*) FROM students st WHERE st.course_id = c.id) AS student_count,
                  (SELECT COALESCE(SUM(p.paid_amount), 0)
                   FROM payments p JOIN students st ON st.id = p.student_id
                   WHERE st.course_id = c.id) AS total_collected,
                  (SELECT COALESCE(SUM({PAYMENT_REMAINING_P_SQL}), 0)
                   FROM payments p JOIN students st ON st.id = p.student_id
                   WHERE st.course_id = c.id) AS total_dues,
                  (SELECT COALESCE(SUM({PAYMENT_TOTAL_P_SQL}), 0)
                   FROM payments p JOIN students st ON st.id = p.student_id
                   WHERE st.course_id = c.id) AS total_expected
           FROM courses c
           LEFT JOIN teachers t ON t.id = c.teacher_id
           ORDER BY c.name"""
    ).fetchall()

    courses_totals = {
        "student_count": sum(int(c["student_count"] or 0) for c in courses),
        "fee": sum(float(c["fee"] or 0) for c in courses),
        "total_collected": sum(float(c["total_collected"] or 0) for c in courses),
        "total_dues": sum(float(c["total_dues"] or 0) for c in courses),
        "total_expected": sum(float(c["total_expected"] or 0) for c in courses),
    }

    today = date.today().isoformat()

    return render_template(
        "dashboard.html",
        total_students=total_students,
        total_teachers=total_teachers,
        total_courses=total_courses,
        collected=collected,
        pending=pending,
        total_expected=total_expected,
        dues=dues,
        dues_total_paid=dues_total_paid,
        dues_total_remaining=dues_total_remaining,
        courses=courses,
        courses_totals=courses_totals,
        today=today,
        dev=dev_context(),
    )


# ----------------------------------------------------------------------
# Teachers
# ----------------------------------------------------------------------
@app.route("/teachers", methods=["GET", "POST"])
@login_required
def teachers():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        phone = request.form.get("phone", "").strip()
        subject = request.form.get("subject", "").strip()
        if not name:
            flash("Teacher name is required.", "error")
        else:
            db.execute(
                "INSERT INTO teachers (name, phone, subject) VALUES (?, ?, ?)",
                (name, phone, subject),
            )
            db.commit()
            flash("Teacher added.", "success")
        return redirect(url_for("teachers"))

    rows = db.execute(
        f"""SELECT t.*,
                  (SELECT COUNT(*) FROM courses c WHERE c.teacher_id = t.id) AS course_count,
                  (SELECT COUNT(*) FROM students st WHERE st.teacher_id = t.id) AS student_count,
                  (SELECT COALESCE(SUM(p.paid_amount), 0)
                   FROM payments p
                   JOIN students s ON s.id = p.student_id
                   WHERE s.teacher_id = t.id) AS total_collected,
                  (SELECT COALESCE(SUM({PAYMENT_REMAINING_P_SQL}), 0)
                   FROM payments p
                   JOIN students s ON s.id = p.student_id
                   WHERE s.teacher_id = t.id) AS dues_amount
           FROM teachers t ORDER BY t.name"""
    ).fetchall()

    teachers_totals = {
        "course_count": sum(int(t["course_count"] or 0) for t in rows),
        "student_count": sum(int(t["student_count"] or 0) for t in rows),
        "total_collected": sum(float(t["total_collected"] or 0) for t in rows),
        "dues_amount": sum(float(t["dues_amount"] or 0) for t in rows),
    }

    return render_template(
        "teachers.html", teachers=rows, teachers_totals=teachers_totals, dev=dev_context()
    )


@app.route("/teachers/delete/<int:teacher_id>", methods=["POST"])
@login_required
def delete_teacher(teacher_id):
    db = get_db()
    db.execute("DELETE FROM teachers WHERE id = ?", (teacher_id,))
    db.commit()
    flash("Teacher removed.", "success")
    return redirect(url_for("teachers"))


@app.route("/teachers/edit/<int:teacher_id>", methods=["GET", "POST"])
@login_required
def edit_teacher(teacher_id):
    db = get_db()
    teacher = db.execute("SELECT * FROM teachers WHERE id = ?", (teacher_id,)).fetchone()
    if not teacher:
        abort(404)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        phone = request.form.get("phone", "").strip()
        subject = request.form.get("subject", "").strip()
        if not name:
            flash("Teacher name is required.", "error")
        else:
            db.execute(
                "UPDATE teachers SET name = ?, phone = ?, subject = ? WHERE id = ?",
                (name, phone, subject, teacher_id),
            )
            db.commit()
            flash("Teacher updated.", "success")
            return redirect(url_for("teachers"))

    return render_template("teacher_edit.html", teacher=teacher, dev=dev_context())


@app.route("/teachers/<int:teacher_id>/dues")
@login_required
def teacher_dues(teacher_id):
    db = get_db()
    teacher = db.execute("SELECT * FROM teachers WHERE id = ?", (teacher_id,)).fetchone()
    if not teacher:
        abort(404)

    dues = db.execute(
        f"""SELECT p.*, s.name AS student_name, s.candidate_no, s.phone,
                  c.name AS course_name,
                  {PAYMENT_TOTAL_P_SQL} AS payment_total,
                  {PAYMENT_REMAINING_P_SQL} AS remaining_amount
           FROM payments p
           JOIN students s ON s.id = p.student_id
           LEFT JOIN courses c ON c.id = s.course_id
           WHERE s.teacher_id = ? AND {PAYMENT_REMAINING_P_SQL} > 0
           ORDER BY s.name, p.installment_no""",
        (teacher_id,),
    ).fetchall()
    total_payable = sum(float(row["payment_total"] or 0) for row in dues)
    total_paid = sum(float(row["paid_amount"] or 0) for row in dues)
    total_dues = sum(float(row["remaining_amount"] or 0) for row in dues)
    return render_template(
        "teacher_dues.html",
        teacher=teacher,
        dues=dues,
        total_payable=total_payable,
        total_paid=total_paid,
        total_dues=total_dues,
        dev=dev_context(),
    )


# ----------------------------------------------------------------------
# Courses
# ----------------------------------------------------------------------
@app.route("/courses", methods=["GET", "POST"])
@login_required
def courses():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        duration = request.form.get("duration", "").strip()
        fee = float(request.form.get("fee") or 0)
        id_card_fee = float(request.form.get("id_card_fee") or 0)
        dmc_fee = float(request.form.get("dmc_fee") or 0)
        exam_fee = float(request.form.get("exam_fee") or 0)
        fund_fee = float(request.form.get("fund_fee") or 0)
        teacher_id = request.form.get("teacher_id") or None
        if not name:
            flash("Course name is required.", "error")
        else:
            db.execute(
                """INSERT INTO courses
                   (name, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee, teacher_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (name, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee, teacher_id),
            )
            db.commit()
            flash("Course added.", "success")
        return redirect(url_for("courses"))

    rows = db.execute(
        f"""SELECT c.*, t.name AS teacher_name,
                  (SELECT COUNT(*) FROM students st WHERE st.course_id = c.id) AS student_count,
                  (SELECT COALESCE(SUM(p.paid_amount), 0)
                   FROM payments p JOIN students st ON st.id = p.student_id
                   WHERE st.course_id = c.id) AS total_collected,
                  (SELECT COALESCE(SUM({PAYMENT_REMAINING_P_SQL}), 0)
                   FROM payments p JOIN students st ON st.id = p.student_id
                   WHERE st.course_id = c.id) AS total_dues,
                  (SELECT COALESCE(SUM({PAYMENT_TOTAL_P_SQL}), 0)
                   FROM payments p JOIN students st ON st.id = p.student_id
                   WHERE st.course_id = c.id) AS total_expected
           FROM courses c
           LEFT JOIN teachers t ON t.id = c.teacher_id
           ORDER BY c.name"""
    ).fetchall()

    courses_totals = {
        "student_count": sum(int(c["student_count"] or 0) for c in rows),
        "fee": sum(float(c["fee"] or 0) for c in rows),
        "total_collected": sum(float(c["total_collected"] or 0) for c in rows),
        "total_dues": sum(float(c["total_dues"] or 0) for c in rows),
        "total_expected": sum(float(c["total_expected"] or 0) for c in rows),
    }

    teacher_rows = db.execute("SELECT * FROM teachers ORDER BY name").fetchall()
    return render_template(
        "courses.html", courses=rows, courses_totals=courses_totals,
        teachers=teacher_rows, dev=dev_context()
    )


@app.route("/courses/delete/<int:course_id>", methods=["POST"])
@login_required
def delete_course(course_id):
    db = get_db()
    db.execute("DELETE FROM courses WHERE id = ?", (course_id,))
    db.commit()
    flash("Course removed.", "success")
    return redirect(url_for("courses"))


@app.route("/courses/edit/<int:course_id>", methods=["GET", "POST"])
@login_required
def edit_course(course_id):
    db = get_db()
    course = db.execute("SELECT * FROM courses WHERE id = ?", (course_id,)).fetchone()
    if not course:
        abort(404)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        duration = request.form.get("duration", "").strip()
        fee = float(request.form.get("fee") or 0)
        id_card_fee = float(request.form.get("id_card_fee") or 0)
        dmc_fee = float(request.form.get("dmc_fee") or 0)
        exam_fee = float(request.form.get("exam_fee") or 0)
        fund_fee = float(request.form.get("fund_fee") or 0)
        teacher_id = request.form.get("teacher_id") or None
        if not name:
            flash("Course name is required.", "error")
        else:
            db.execute(
                """UPDATE courses
                   SET name = ?, duration = ?, fee = ?, id_card_fee = ?, dmc_fee = ?,
                       exam_fee = ?, fund_fee = ?, teacher_id = ?
                   WHERE id = ?""",
                (name, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee, teacher_id, course_id),
            )
            db.commit()
            flash("Course updated. Existing students keep their current fee schedule unless edited separately.", "success")
            return redirect(url_for("courses"))

    teacher_rows = db.execute("SELECT * FROM teachers ORDER BY name").fetchall()
    return render_template("course_edit.html", course=course, teachers=teacher_rows, dev=dev_context())


# ----------------------------------------------------------------------
# Students
# ----------------------------------------------------------------------
@app.route("/students", methods=["GET", "POST"])
@login_required
def students():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        father_name = request.form.get("father_name", "").strip()
        phone = request.form.get("phone", "").strip()
        candidate_no = request.form.get("candidate_no", "").strip()
        course_id = request.form.get("course_id") or None
        teacher_id = request.form.get("teacher_id") or None
        total_fee = float(request.form.get("total_fee") or 0)
        installment_count = int(request.form.get("installment_count") or 1)
        admission_date = request.form.get("admission_date") or date.today().isoformat()

        if not name:
            flash("Student name is required.", "error")
            return redirect(url_for("students"))

        cur = db.execute(
            """INSERT INTO students
               (candidate_no, name, father_name, phone, course_id, teacher_id,
                total_fee, installment_count, admission_date)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (candidate_no, name, father_name, phone, course_id, teacher_id,
             total_fee, installment_count, admission_date),
        )
        db.commit()
        student_id = cur.lastrowid

        course = None
        if course_id:
            course = db.execute("SELECT * FROM courses WHERE id = ?", (course_id,)).fetchone()

        create_payments_for_student(db, student_id, total_fee, installment_count, course, admission_date)
        flash("Student enrolled and fee schedule generated.", "success")
        return redirect(url_for("view_student", student_id=student_id))

    rows = db.execute(
        f"""SELECT s.*, c.name AS course_name, t.name AS teacher_name,
                  (SELECT COUNT(*) FROM payments p
                   WHERE p.student_id = s.id AND {PAYMENT_REMAINING_P_SQL} > 0) AS pending_installments,
                  (SELECT COALESCE(SUM(p.paid_amount), 0) FROM payments p
                   WHERE p.student_id = s.id) AS total_paid,
                  (SELECT COALESCE(SUM({PAYMENT_TOTAL_P_SQL}), 0) FROM payments p
                   WHERE p.student_id = s.id) AS total_payable,
                  (SELECT COALESCE(SUM({PAYMENT_REMAINING_P_SQL}), 0) FROM payments p
                   WHERE p.student_id = s.id) AS dues_amount
           FROM students s
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           ORDER BY s.created_at DESC"""
    ).fetchall()

    students_totals = {
        "student_count": len(rows),
        "total_fee": sum(float(s["total_fee"] or 0) for s in rows),
        "total_payable": sum(float(s["total_payable"] or 0) for s in rows),
        "total_paid": sum(float(s["total_paid"] or 0) for s in rows),
        "dues_amount": sum(float(s["dues_amount"] or 0) for s in rows),
    }

    course_rows = db.execute("SELECT * FROM courses ORDER BY name").fetchall()
    teacher_rows = db.execute("SELECT * FROM teachers ORDER BY name").fetchall()
    return render_template(
        "students.html", students=rows, students_totals=students_totals,
        courses=course_rows, teachers=teacher_rows,
        today=date.today().isoformat(), dev=dev_context(),
    )


@app.route("/students/<int:student_id>")
@login_required
def view_student(student_id):
    db = get_db()
    student = db.execute(
        """SELECT s.*, c.name AS course_name, c.duration AS course_duration, t.name AS teacher_name
           FROM students s
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           WHERE s.id = ?""",
        (student_id,),
    ).fetchone()
    if not student:
        abort(404)
    payments = get_student_payment_rows(db, student_id)
    total_payable, total_paid, total_due = summarize_payments(payments)
    return render_template(
        "student_view.html",
        student=student,
        payments=payments,
        total_payable=total_payable,
        total_paid=total_paid,
        total_due=total_due,
        dev=dev_context(),
    )


@app.route("/students/<int:student_id>/clearance")
@login_required
def student_clearance(student_id):
    db = get_db()
    student = db.execute(
        """SELECT s.*, c.name AS course_name, c.duration AS course_duration, t.name AS teacher_name
           FROM students s
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           WHERE s.id = ?""",
        (student_id,),
    ).fetchone()
    if not student:
        abort(404)

    payments = get_student_payment_rows(db, student_id)
    total_payable, total_paid, total_due = summarize_payments(payments)
    wa_number = format_pk_whatsapp(student["phone"])
    history_lines = "\n".join(
        f"Installment {p['installment_no']}: Payable PKR {float(p['payment_total'] or 0):.0f}, "
        f"Paid PKR {float(p['paid_amount'] or 0):.0f}, "
        f"Remaining PKR {float(p['remaining_amount'] or 0):.0f}"
        for p in payments
    )
    wa_message = (
        f"Fee Clearance Slip - {COLLEGE_NAME}\n"
        f"Student: {student['name']}\n"
        f"Candidate No: {student['candidate_no'] or '-'}\n"
        f"Program: {student['course_name'] or '-'}\n"
        f"Total payable: PKR {total_payable:.0f}\n"
        f"Total paid: PKR {total_paid:.0f}\n"
        f"Remaining dues: PKR {total_due:.0f}\n"
        f"Status: {'CLEARED' if total_due <= 0 else 'DUES REMAINING'}\n\n"
        f"Payment History:\n{history_lines}"
    )
    return render_template(
        "clearance.html",
        student=student,
        payments=payments,
        total_payable=total_payable,
        total_paid=total_paid,
        total_due=total_due,
        wa_number=wa_number,
        wa_message=wa_message,
        today=date.today().strftime("%d/%m/%Y"),
        dev=dev_context(),
    )


@app.route("/students/<int:student_id>/clearance/pdf")
@login_required
def student_clearance_pdf(student_id):
    from pdf_generator import build_clearance_pdf

    db = get_db()
    student = db.execute(
        """SELECT s.*, c.name AS course_name, c.duration AS course_duration, t.name AS teacher_name
           FROM students s
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           WHERE s.id = ?""",
        (student_id,),
    ).fetchone()
    if not student:
        abort(404)

    payments = get_student_payment_rows(db, student_id)
    total_payable, total_paid, total_due = summarize_payments(payments)
    buffer = build_clearance_pdf(
        student,
        payments,
        total_payable,
        total_paid,
        total_due,
        dev_context(),
        COLLEGE_NAME,
    )
    student_name = safe_filename_part(student["name"])
    candidate_no = safe_filename_part(student["candidate_no"] or student["id"])
    filename = f"Clearance_{student_name}_{candidate_no}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


@app.route("/students/delete/<int:student_id>", methods=["POST"])
@login_required
def delete_student(student_id):
    db = get_db()
    db.execute("DELETE FROM students WHERE id = ?", (student_id,))
    db.commit()
    flash("Student removed.", "success")
    return redirect(url_for("students"))


@app.route("/students/edit/<int:student_id>", methods=["GET", "POST"])
@login_required
def edit_student(student_id):
    db = get_db()
    student = db.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
    if not student:
        abort(404)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        father_name = request.form.get("father_name", "").strip()
        phone = request.form.get("phone", "").strip()
        candidate_no = request.form.get("candidate_no", "").strip()
        course_id = request.form.get("course_id") or None
        teacher_id = request.form.get("teacher_id") or None
        total_fee = float(request.form.get("total_fee") or 0)
        installment_count = int(request.form.get("installment_count") or 1)
        admission_date = request.form.get("admission_date") or date.today().isoformat()

        if not name:
            flash("Student name is required.", "error")
            return redirect(url_for("edit_student", student_id=student_id))

        schedule_changed = (
            str(student["course_id"] or "") != str(course_id or "")
            or float(student["total_fee"] or 0) != total_fee
            or int(student["installment_count"] or 1) != installment_count
            or (student["admission_date"] or "") != admission_date
        )

        paid_max = db.execute(
            "SELECT MAX(installment_no) AS max_paid FROM payments WHERE student_id = ? AND (paid = 1 OR paid_amount > 0)",
            (student_id,),
        ).fetchone()["max_paid"]
        if paid_max and installment_count < paid_max:
            installment_count = paid_max
            flash(f"Installments kept at {paid_max} because paid voucher records already exist.", "error")

        db.execute(
            """UPDATE students
               SET candidate_no = ?, name = ?, father_name = ?, phone = ?, course_id = ?,
                   teacher_id = ?, total_fee = ?, installment_count = ?, admission_date = ?
               WHERE id = ?""",
            (candidate_no, name, father_name, phone, course_id, teacher_id,
             total_fee, installment_count, admission_date, student_id),
        )

        if schedule_changed:
            course = None
            if course_id:
                course = db.execute("SELECT * FROM courses WHERE id = ?", (course_id,)).fetchone()
            rebuild_unpaid_payments_for_student(
                db, student_id, total_fee, installment_count, course, admission_date
            )
        else:
            db.commit()

        flash("Student updated.", "success")
        return redirect(url_for("view_student", student_id=student_id))

    course_rows = db.execute("SELECT * FROM courses ORDER BY name").fetchall()
    teacher_rows = db.execute("SELECT * FROM teachers ORDER BY name").fetchall()
    return render_template(
        "student_edit.html",
        student=student,
        courses=course_rows,
        teachers=teacher_rows,
        dev=dev_context(),
    )


@app.route("/payments/<int:payment_id>/mark_paid", methods=["POST"])
@login_required
def mark_paid(payment_id):
    db = get_db()
    payment = db.execute("SELECT * FROM payments WHERE id = ?", (payment_id,)).fetchone()
    if not payment:
        abort(404)
    total = payment_total(payment)
    current_paid = float(payment["paid_amount"] or 0)
    remaining = max(total - current_paid, 0)
    requested_amount = request.form.get("paid_amount")
    try:
        amount = float(requested_amount) if requested_amount else remaining
    except ValueError:
        flash("Please enter a valid payment amount.", "error")
        return redirect(request.referrer or url_for("voucher", payment_id=payment_id))

    if amount <= 0:
        flash("Payment amount must be greater than zero.", "error")
        return redirect(request.referrer or url_for("voucher", payment_id=payment_id))

    new_paid_amount = min(current_paid + amount, total)
    is_paid = 1 if new_paid_amount >= total else 0
    db.execute(
        "UPDATE payments SET paid_amount = ?, paid = ?, paid_date = ? WHERE id = ?",
        (new_paid_amount, is_paid, date.today().isoformat(), payment_id),
    )
    db.commit()
    if is_paid:
        flash("Installment fully paid.", "success")
    else:
        flash(f"Partial payment saved. Remaining: Rs. {total - new_paid_amount:.0f}", "success")
    return redirect(request.referrer or url_for("voucher", payment_id=payment_id))


# ----------------------------------------------------------------------
# Voucher / Slip
# ----------------------------------------------------------------------
def get_voucher_data(db, payment_id):
    payment = db.execute(
        f"""SELECT *,
                  {PAYMENT_TOTAL_P_SQL} AS payment_total,
                  {PAYMENT_REMAINING_P_SQL} AS remaining_amount
           FROM payments p
           WHERE id = ?""",
        (payment_id,),
    ).fetchone()
    if not payment:
        return None, None
    student = db.execute(
        """SELECT s.*, c.name AS course_name, c.duration AS course_duration, t.name AS teacher_name
           FROM students s
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           WHERE s.id = ?""",
        (payment["student_id"],),
    ).fetchone()
    return payment, student


@app.route("/voucher/<int:payment_id>")
@login_required
def voucher(payment_id):
    db = get_db()
    payment, student = get_voucher_data(db, payment_id)
    if not payment or not student:
        abort(404)
    total = float(payment["payment_total"] or payment_total(payment))
    paid_amount = float(payment["paid_amount"] or 0)
    remaining_amount = float(payment["remaining_amount"] or payment_remaining(payment))
    student_summary = db.execute(
        f"""SELECT COALESCE(SUM(paid_amount), 0) AS total_paid,
                  COALESCE(SUM({PAYMENT_REMAINING_P_SQL}), 0) AS total_due
           FROM payments p
           WHERE student_id = ?""",
        (student["id"],),
    ).fetchone()
    student_total_paid = float(student_summary["total_paid"] or 0)
    student_total_due = float(student_summary["total_due"] or 0)
    wa_number = format_pk_whatsapp(student["phone"])
    wa_message = (
        f"Fee Voucher - {COLLEGE_NAME}\n"
        f"Student: {student['name']}\n"
        f"Candidate No: {student['candidate_no'] or '-'}\n"
        f"Installment: {payment['installment_no']} of {student['installment_count']}\n"
        f"Amount: PKR {total:.0f}\n"
        f"Paid on this installment: PKR {paid_amount:.0f}\n"
        f"Remaining on this installment: PKR {remaining_amount:.0f}\n"
        f"Student total remaining dues: PKR {student_total_due:.0f}\n"
        f"Due Date: {payment['due_date']}\n"
        f"Please find the attached fee voucher PDF."
    )
    return render_template(
        "voucher.html",
        payment=payment,
        student=student,
        total=total,
        paid_amount=paid_amount,
        remaining_amount=remaining_amount,
        student_total_paid=student_total_paid,
        student_total_due=student_total_due,
        wa_number=wa_number,
        wa_message=wa_message,
        today=date.today().strftime("%d/%m/%Y"),
        dev=dev_context(),
    )


@app.route("/voucher/<int:payment_id>/pdf")
@login_required
def voucher_pdf(payment_id):
    from pdf_generator import build_voucher_pdf

    db = get_db()
    payment, student = get_voucher_data(db, payment_id)
    if not payment or not student:
        abort(404)

    student_summary = db.execute(
        f"""SELECT COALESCE(SUM(paid_amount), 0) AS total_paid,
                  COALESCE(SUM({PAYMENT_REMAINING_P_SQL}), 0) AS total_due
           FROM payments p
           WHERE student_id = ?""",
        (student["id"],),
    ).fetchone()
    buffer = build_voucher_pdf(
        payment,
        student,
        dev_context(),
        COLLEGE_NAME,
        float(student_summary["total_due"] or 0),
    )
    student_name = safe_filename_part(student["name"])
    candidate_no = safe_filename_part(student["candidate_no"] or student["id"])
    filename = f"Voucher_{student_name}_{candidate_no}_Inst{payment['installment_no']}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


# ----------------------------------------------------------------------
# Settings (change admin password)
# ----------------------------------------------------------------------
@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    db = get_db()
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        admin = db.execute("SELECT * FROM admin WHERE id = ?", (session["admin_id"],)).fetchone()
        if not check_password_hash(admin["password_hash"], current_password):
            flash("Current password is incorrect.", "error")
        elif len(new_password) < 6:
            flash("New password must be at least 6 characters.", "error")
        elif new_password != confirm_password:
            flash("New passwords do not match.", "error")
        else:
            db.execute(
                "UPDATE admin SET password_hash = ? WHERE id = ?",
                (generate_password_hash(new_password), session["admin_id"]),
            )
            db.commit()
            flash("Password updated successfully.", "success")
        return redirect(url_for("settings"))

    return render_template("settings.html", dev=dev_context())


# ----------------------------------------------------------------------
# Excel Export / Import
# ----------------------------------------------------------------------
@app.route("/data")
@app.route("/data-tools")
@app.route("/data_tools")
@login_required
def data_tools():
    return render_template("data_tools.html", dev=dev_context())


@app.route("/export/excel")
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = get_db()
    wb = openpyxl.Workbook()

    primary_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
    dark_fill = PatternFill(start_color="0B4CB8", end_color="0B4CB8", fill_type="solid")
    total_fill = PatternFill(start_color="EEF3FC", end_color="EEF3FC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(color="FFFFFF", bold=True, size=14)
    bold_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE")
    )
    total_border = Border(
        top=Side(style="thin", color="1F6FEB"),
        bottom=Side(style="double", color="1F6FEB")
    )

    def style_table(ws, start_row, ncols, has_total=True):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = primary_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        last_row = ws.max_row
        if has_total and last_row > start_row:
            for col in range(1, ncols + 1):
                cell = ws.cell(row=last_row, column=col)
                cell.fill = total_fill
                cell.font = bold_font
                cell.border = total_border

    # 1. Summary Sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 34

    ws_sum.merge_cells("A1:B1")
    title_cell = ws_sum["A1"]
    title_cell.value = ACADEMY_NAME
    title_cell.fill = dark_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 32

    ws_sum.merge_cells("A2:B2")
    sub_cell = ws_sum["A2"]
    sub_cell.value = "Fee Voucher & Financial Management Report"
    sub_cell.font = Font(italic=True, size=10, color="555555")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    total_students = db.execute("SELECT COUNT(*) c FROM students").fetchone()["c"]
    total_teachers = db.execute("SELECT COUNT(*) c FROM teachers").fetchone()["c"]
    total_courses = db.execute("SELECT COUNT(*) c FROM courses").fetchone()["c"]
    total_collected = db.execute("SELECT COALESCE(SUM(paid_amount), 0) s FROM payments").fetchone()["s"]
    total_pending = db.execute(f"SELECT COALESCE(SUM({PAYMENT_REMAINING_SQL}), 0) s FROM payments").fetchone()["s"]
    total_revenue = total_collected + total_pending
    recovery_rate = (total_collected / total_revenue * 100) if total_revenue > 0 else 0

    metrics = [
        ("Report Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Institution Name", ACADEMY_NAME),
        ("Admin / Support Contact", f"{ADMIN_NAME} ({ADMIN_CONTACT})"),
        ("Total Active Students", total_students),
        ("Total Teachers", total_teachers),
        ("Total Courses", total_courses),
        ("Total Revenue Expected (PKR)", f"{total_revenue:,.0f}"),
        ("Total Amount Collected (PKR)", f"{total_collected:,.0f}"),
        ("Total Pending Dues (PKR)", f"{total_pending:,.0f}"),
        ("Recovery Rate (%)", f"{recovery_rate:.1f}%"),
    ]

    for idx, (label, val) in enumerate(metrics, start=4):
        ws_sum[f"A{idx}"] = label
        ws_sum[f"B{idx}"] = val
        ws_sum[f"A{idx}"].font = bold_font
        ws_sum[f"A{idx}"].border = thin_border
        ws_sum[f"B{idx}"].border = thin_border

    # 2. Courses Sheet
    ws_c = wb.create_sheet("Courses")
    headers_c = [
        "Course ID", "Course Name", "Teacher Name", "Duration", "Tuition Fee (PKR)",
        "ID Card Fee", "DMC Fee", "Exam Fee", "Fund Fee", "Total Course Fee",
        "Enrolled Students", "Total Collected (PKR)", "Total Remaining Dues (PKR)", "Total Expected Revenue (PKR)"
    ]
    ws_c.append(headers_c)
    courses_data = db.execute(
        f"""SELECT c.*, t.name AS teacher_name,
                  (SELECT COUNT(*) FROM students st WHERE st.course_id = c.id) AS student_count,
                  (SELECT COALESCE(SUM(p.paid_amount), 0)
                   FROM payments p JOIN students st ON st.id = p.student_id
                   WHERE st.course_id = c.id) AS total_collected,
                  (SELECT COALESCE(SUM({PAYMENT_REMAINING_P_SQL}), 0)
                   FROM payments p JOIN students st ON st.id = p.student_id
                   WHERE st.course_id = c.id) AS total_dues,
                  (SELECT COALESCE(SUM({PAYMENT_TOTAL_P_SQL}), 0)
                   FROM payments p JOIN students st ON st.id = p.student_id
                   WHERE st.course_id = c.id) AS total_expected
           FROM courses c
           LEFT JOIN teachers t ON t.id = c.teacher_id
           ORDER BY c.id"""
    ).fetchall()

    sum_c_students = sum_c_fee = sum_c_col = sum_c_dues = sum_c_exp = 0
    for c in courses_data:
        tot_course_fee = float(c["fee"] or 0) + float(c["id_card_fee"] or 0) + float(c["dmc_fee"] or 0) + float(c["exam_fee"] or 0) + float(c["fund_fee"] or 0)
        st_count = int(c["student_count"] or 0)
        col_amt = float(c["total_collected"] or 0)
        dues_amt = float(c["total_dues"] or 0)
        exp_amt = float(c["total_expected"] or 0)

        sum_c_students += st_count
        sum_c_fee += float(c["fee"] or 0)
        sum_c_col += col_amt
        sum_c_dues += dues_amt
        sum_c_exp += exp_amt

        ws_c.append([
            c["id"], c["name"], c["teacher_name"] or "Unassigned", c["duration"] or "-",
            float(c["fee"] or 0), float(c["id_card_fee"] or 0), float(c["dmc_fee"] or 0),
            float(c["exam_fee"] or 0), float(c["fund_fee"] or 0), tot_course_fee,
            st_count, col_amt, dues_amt, exp_amt
        ])
    ws_c.append(["TOTAL", "", "", "", sum_c_fee, "", "", "", "", "", sum_c_students, sum_c_col, sum_c_dues, sum_c_exp])
    style_table(ws_c, 1, len(headers_c))

    # 3. Students Sheet
    ws_s = wb.create_sheet("Students")
    headers_s = [
        "Student ID", "Candidate No", "Student Name", "Father Name", "Phone / WhatsApp",
        "Course Name", "Teacher Name", "Admission Date", "Installments",
        "Total Course Fee (PKR)", "Total Payable (PKR)", "Total Paid (PKR)", "Remaining Dues (PKR)", "Status"
    ]
    ws_s.append(headers_s)
    students_data = db.execute(
        f"""SELECT s.*, c.name AS course_name, t.name AS teacher_name,
                  (SELECT COALESCE(SUM(p.paid_amount), 0) FROM payments p WHERE p.student_id = s.id) AS total_paid,
                  (SELECT COALESCE(SUM({PAYMENT_TOTAL_P_SQL}), 0) FROM payments p WHERE p.student_id = s.id) AS total_payable,
                  (SELECT COALESCE(SUM({PAYMENT_REMAINING_P_SQL}), 0) FROM payments p WHERE p.student_id = s.id) AS dues_amount
           FROM students s
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           ORDER BY s.id"""
    ).fetchall()

    sum_s_fee = sum_s_pay = sum_s_paid = sum_s_dues = 0
    for s in students_data:
        fee = float(s["total_fee"] or 0)
        pay = float(s["total_payable"] or 0)
        paid = float(s["total_paid"] or 0)
        dues = float(s["dues_amount"] or 0)
        status = "CLEARED" if dues <= 0 else "PENDING DUES"

        sum_s_fee += fee
        sum_s_pay += pay
        sum_s_paid += paid
        sum_s_dues += dues

        ws_s.append([
            s["id"], s["candidate_no"] or "-", s["name"], s["father_name"] or "-", s["phone"] or "-",
            s["course_name"] or "Unassigned", s["teacher_name"] or "Unassigned", s["admission_date"] or "-",
            s["installment_count"], fee, pay, paid, dues, status
        ])
    ws_s.append(["TOTAL", f"Total: {len(students_data)} Students", "", "", "", "", "", "", "", sum_s_fee, sum_s_pay, sum_s_paid, sum_s_dues, ""])
    style_table(ws_s, 1, len(headers_s))

    # 4. Payments / Installments Sheet
    ws_p = wb.create_sheet("Payments")
    headers_p = [
        "Payment ID", "Student ID", "Candidate No", "Student Name", "Father Name",
        "Course Name", "Teacher Name", "Installment No", "Due Date",
        "Tuition (PKR)", "ID Card Fee", "DMC Fee", "Exam Fee", "Fund Fee",
        "Installment Payable (PKR)", "Amount Paid (PKR)", "Remaining Dues (PKR)", "Status", "Paid Date"
    ]
    ws_p.append(headers_p)
    payments_data = db.execute(
        f"""SELECT p.*, s.name AS student_name, s.father_name, s.candidate_no,
                  c.name AS course_name, t.name AS teacher_name,
                  {PAYMENT_TOTAL_P_SQL} AS payment_total,
                  {PAYMENT_REMAINING_P_SQL} AS remaining_amount
           FROM payments p
           JOIN students s ON s.id = p.student_id
           LEFT JOIN courses c ON c.id = s.course_id
           LEFT JOIN teachers t ON t.id = s.teacher_id
           ORDER BY p.student_id, p.installment_no"""
    ).fetchall()

    sum_p_tuition = sum_p_pay = sum_p_paid = sum_p_rem = 0
    for p in payments_data:
        tuition = float(p["tuition_amount"] or 0)
        pay = float(p["payment_total"] or 0)
        paid = float(p["paid_amount"] or 0)
        rem = float(p["remaining_amount"] or 0)
        status = "PAID" if p["paid"] else ("PARTIAL" if paid > 0 else "UNPAID")

        sum_p_tuition += tuition
        sum_p_pay += pay
        sum_p_paid += paid
        sum_p_rem += rem

        ws_p.append([
            p["id"], p["student_id"], p["candidate_no"] or "-", p["student_name"], p["father_name"] or "-",
            p["course_name"] or "-", p["teacher_name"] or "-", p["installment_no"], p["due_date"] or "-",
            tuition, float(p["id_card_fee"] or 0), float(p["dmc_fee"] or 0),
            float(p["exam_fee"] or 0), float(p["fund_fee"] or 0),
            pay, paid, rem, status, p["paid_date"] or "-"
        ])
    ws_p.append(["TOTAL", f"Total: {len(payments_data)} Installments", "", "", "", "", "", "", "", sum_p_tuition, "", "", "", "", sum_p_pay, sum_p_paid, sum_p_rem, "", ""])
    style_table(ws_p, 1, len(headers_p))

    # 5. Teachers Sheet
    ws_t = wb.create_sheet("Teachers")
    headers_t = [
        "Teacher ID", "Teacher Name", "Phone", "Subject",
        "Assigned Courses", "Enrolled Students", "Total Amount Collected (PKR)", "Total Pending Dues (PKR)"
    ]
    ws_t.append(headers_t)
    teachers_data = db.execute(
        f"""SELECT t.*,
                  (SELECT COUNT(*) FROM courses c WHERE c.teacher_id = t.id) AS course_count,
                  (SELECT COUNT(*) FROM students st WHERE st.teacher_id = t.id) AS student_count,
                  (SELECT COALESCE(SUM(p.paid_amount), 0)
                   FROM payments p JOIN students s ON s.id = p.student_id
                   WHERE s.teacher_id = t.id) AS total_collected,
                  (SELECT COALESCE(SUM({PAYMENT_REMAINING_P_SQL}), 0)
                   FROM payments p JOIN students s ON s.id = p.student_id
                   WHERE s.teacher_id = t.id) AS dues_amount
           FROM teachers t ORDER BY t.id"""
    ).fetchall()

    sum_t_courses = sum_t_students = sum_t_col = sum_t_dues = 0
    for t in teachers_data:
        c_cnt = int(t["course_count"] or 0)
        s_cnt = int(t["student_count"] or 0)
        col = float(t["total_collected"] or 0)
        dues = float(t["dues_amount"] or 0)

        sum_t_courses += c_cnt
        sum_t_students += s_cnt
        sum_t_col += col
        sum_t_dues += dues

        ws_t.append([
            t["id"], t["name"], t["phone"] or "-", t["subject"] or "-",
            c_cnt, s_cnt, col, dues
        ])
    ws_t.append(["TOTAL", f"Total: {len(teachers_data)} Teachers", "", "", sum_t_courses, sum_t_students, sum_t_col, sum_t_dues])
    style_table(ws_t, 1, len(headers_t))

    # Auto-adjust column widths across all sheets
    from openpyxl.utils import get_column_letter
    for sheet in wb.worksheets:
        if sheet.title == "Summary":
            sheet.column_dimensions["A"].width = 30
            sheet.column_dimensions["B"].width = 36
            continue
        for col_idx, col_cells in enumerate(sheet.columns, start=1):
            col_letter = get_column_letter(col_idx)
            length = max(len(str(c.value or "")) for c in col_cells)
            sheet.column_dimensions[col_letter].width = min(max(length + 3, 12), 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"ssaq_full_details_export_{date.today().isoformat()}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/import/excel", methods=["POST"])
@login_required
def import_excel():
    import openpyxl

    file = request.files.get("excel_file")
    if not file or file.filename == "":
        flash("Please choose an .xlsx file to import.", "error")
        return redirect(url_for("data_tools"))
    if not file.filename.lower().endswith(".xlsx"):
        flash("Only .xlsx files are supported.", "error")
        return redirect(url_for("data_tools"))

    db = get_db()
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        flash(f"Could not read the Excel file: {e}", "error")
        return redirect(url_for("data_tools"))

    teacher_map = {}       # old_id -> new_id / teacher_name -> new_id
    course_map = {}        # old_id -> new_id / course_name -> new_id
    student_map = {}       # old_id -> new_id
    student_cand_map = {}  # candidate_no.lower() -> new_id
    student_name_map = {}  # (name.lower(), father_name.lower()) -> new_id and name.lower() -> new_id
    counts = {"teachers": 0, "courses": 0, "students": 0, "payments": 0}

    # 1. Teachers
    if "Teachers" in wb.sheetnames:
        ws = wb["Teachers"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[1] or str(row[0]).strip().upper() == "TOTAL":
                continue
            old_id, name, phone, subject = (list(row) + [None] * 4)[:4]
            name = str(name).strip()
            existing = db.execute("SELECT id FROM teachers WHERE LOWER(name) = LOWER(?)", (name,)).fetchone()
            if existing:
                t_id = existing["id"]
                db.execute("UPDATE teachers SET phone = ?, subject = ? WHERE id = ?",
                           (str(phone or "").strip(), str(subject or "").strip(), t_id))
            else:
                cur = db.execute(
                    "INSERT INTO teachers (name, phone, subject) VALUES (?, ?, ?)",
                    (name, str(phone or "").strip(), str(subject or "").strip()),
                )
                t_id = cur.lastrowid
                counts["teachers"] += 1
            if old_id is not None:
                try:
                    teacher_map[int(old_id)] = t_id
                except (ValueError, TypeError):
                    pass
            teacher_map[name.lower()] = t_id
        db.commit()

    # 2. Courses
    if "Courses" in wb.sheetnames:
        ws = wb["Courses"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[1] or str(row[0]).strip().upper() == "TOTAL":
                continue
            padded = (list(row) + [None] * 10)[:10]
            old_id, name, teacher_ref, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee = padded[:9]
            name = str(name).strip()

            resolved_teacher_id = None
            if teacher_ref not in (None, "", "Unassigned", "-"):
                try:
                    resolved_teacher_id = teacher_map.get(int(teacher_ref))
                except (ValueError, TypeError):
                    resolved_teacher_id = teacher_map.get(str(teacher_ref).strip().lower())

            existing = db.execute("SELECT id FROM courses WHERE LOWER(name) = LOWER(?)", (name,)).fetchone()
            if existing:
                c_id = existing["id"]
                db.execute(
                    """UPDATE courses SET duration = ?, fee = ?, id_card_fee = ?, dmc_fee = ?,
                                          exam_fee = ?, fund_fee = ?, teacher_id = COALESCE(?, teacher_id)
                       WHERE id = ?""",
                    (str(duration or "").strip(), float(fee or 0), float(id_card_fee or 0),
                     float(dmc_fee or 0), float(exam_fee or 0), float(fund_fee or 0), resolved_teacher_id, c_id)
                )
            else:
                cur = db.execute(
                    """INSERT INTO courses (name, duration, fee, id_card_fee, dmc_fee, exam_fee, fund_fee, teacher_id)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (name, str(duration or "").strip(), float(fee or 0), float(id_card_fee or 0),
                     float(dmc_fee or 0), float(exam_fee or 0), float(fund_fee or 0), resolved_teacher_id),
                )
                c_id = cur.lastrowid
                counts["courses"] += 1
            if old_id is not None:
                try:
                    course_map[int(old_id)] = c_id
                except (ValueError, TypeError):
                    pass
            course_map[name.lower()] = c_id
        db.commit()

    # 3. Students
    has_payments_sheet = "Payments" in wb.sheetnames
    if "Students" in wb.sheetnames:
        ws = wb["Students"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[2] or str(row[0]).strip().upper() == "TOTAL":
                continue
            padded = (list(row) + [None] * 14)[:14]
            (old_id, candidate_no, name, father_name, phone, course_ref,
             teacher_ref, admission_date, installment_count, total_fee, total_payable, paid_amount) = padded[:12]

            name = str(name).strip()
            father_name_str = str(father_name or "").strip()
            candidate_no_str = str(candidate_no or "").strip()
            if candidate_no_str == "-":
                candidate_no_str = ""

            resolved_course_id = None
            if course_ref not in (None, "", "Unassigned", "-"):
                try:
                    resolved_course_id = course_map.get(int(course_ref))
                except (ValueError, TypeError):
                    resolved_course_id = course_map.get(str(course_ref).strip().lower())

            resolved_teacher_id = None
            if teacher_ref not in (None, "", "Unassigned", "-"):
                try:
                    resolved_teacher_id = teacher_map.get(int(teacher_ref))
                except (ValueError, TypeError):
                    resolved_teacher_id = teacher_map.get(str(teacher_ref).strip().lower())

            admission_date_str = str(admission_date)[:10] if admission_date and str(admission_date) != "-" else date.today().isoformat()

            # Find or insert student
            existing_student = None
            if candidate_no_str:
                existing_student = db.execute("SELECT id FROM students WHERE candidate_no = ?", (candidate_no_str,)).fetchone()
            if not existing_student and name:
                existing_student = db.execute(
                    "SELECT id FROM students WHERE LOWER(name) = LOWER(?) AND LOWER(father_name) = LOWER(?)",
                    (name, father_name_str)
                ).fetchone()

            if existing_student:
                s_id = existing_student["id"]
                db.execute(
                    """UPDATE students SET candidate_no = ?, name = ?, father_name = ?, phone = ?,
                                          course_id = COALESCE(?, course_id),
                                          teacher_id = COALESCE(?, teacher_id),
                                          total_fee = ?, installment_count = ?, admission_date = ?
                       WHERE id = ?""",
                    (candidate_no_str, name, father_name_str, str(phone or "").strip(),
                     resolved_course_id, resolved_teacher_id, float(total_fee or 0),
                     int(installment_count or 1), admission_date_str, s_id)
                )
            else:
                cur = db.execute(
                    """INSERT INTO students (candidate_no, name, father_name, phone, course_id, teacher_id,
                                              total_fee, installment_count, admission_date)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (candidate_no_str, name, father_name_str, str(phone or "").strip(),
                     resolved_course_id, resolved_teacher_id, float(total_fee or 0),
                     int(installment_count or 1), admission_date_str),
                )
                s_id = cur.lastrowid
                counts["students"] += 1

            if old_id is not None:
                try:
                    student_map[int(old_id)] = s_id
                except (ValueError, TypeError):
                    pass
            if candidate_no_str:
                student_cand_map[candidate_no_str.lower()] = s_id
            student_name_map[(name.lower(), father_name_str.lower())] = s_id
            student_name_map[name.lower()] = s_id

            # If Payments sheet is NOT present, auto-create payment schedule and apply paid amount if present
            if not has_payments_sheet:
                db.execute("DELETE FROM payments WHERE student_id = ?", (s_id,))
                course = None
                if resolved_course_id:
                    course = db.execute("SELECT * FROM courses WHERE id = ?", (resolved_course_id,)).fetchone()
                create_payments_for_student(
                    db, s_id, float(total_fee or 0), int(installment_count or 1), course, admission_date_str
                )
                try:
                    p_amt = float(paid_amount or 0)
                    if p_amt > 0:
                        payments = db.execute("SELECT * FROM payments WHERE student_id = ? ORDER BY installment_no", (s_id,)).fetchall()
                        rem_to_apply = p_amt
                        for p in payments:
                            p_tot = float(p["tuition_amount"]) + float(p["id_card_fee"]) + float(p["dmc_fee"]) + float(p["exam_fee"]) + float(p["fund_fee"])
                            if rem_to_apply >= p_tot:
                                db.execute("UPDATE payments SET paid = 1, paid_amount = ?, paid_date = ? WHERE id = ?",
                                           (p_tot, admission_date_str, p["id"]))
                                rem_to_apply -= p_tot
                            elif rem_to_apply > 0:
                                db.execute("UPDATE payments SET paid = 0, paid_amount = ? WHERE id = ?",
                                           (rem_to_apply, p["id"]))
                                rem_to_apply = 0
                except (ValueError, TypeError):
                    pass

        db.commit()

    # 4. Payments Sheet (Exact restoration of all installments, collections, paid amounts, dates, and statuses)
    if has_payments_sheet:
        ws_p = wb["Payments"]
        rows_p = list(ws_p.iter_rows(min_row=2, values_only=True))

        # Clear existing payments for imported students before inserting exact payments from Payments sheet
        for s_id in set(student_map.values()):
            db.execute("DELETE FROM payments WHERE student_id = ?", (s_id,))

        for row in rows_p:
            if not row or str(row[0]).strip().upper() == "TOTAL":
                continue
            padded_p = (list(row) + [None] * 19)[:19]
            (old_p_id, old_s_id, cand_no, s_name, f_name, c_name, t_name,
             inst_no, due_date, tuition, id_fee, dmc_fee, exam_fee, fund_fee,
             payable, paid_amt, rem_dues, status, paid_date) = padded_p

            target_student_id = None
            if old_s_id is not None:
                try:
                    target_student_id = student_map.get(int(old_s_id))
                except (ValueError, TypeError):
                    pass
            if not target_student_id and cand_no and str(cand_no).strip() != "-":
                target_student_id = student_cand_map.get(str(cand_no).strip().lower())
            if not target_student_id and s_name:
                target_student_id = student_name_map.get((str(s_name).strip().lower(), str(f_name or "").strip().lower()))
            if not target_student_id and s_name:
                target_student_id = student_name_map.get(str(s_name).strip().lower())

            if target_student_id:
                inst_no_val = int(inst_no or 1)
                due_date_str = str(due_date)[:10] if due_date and str(due_date).strip() != "-" else date.today().isoformat()
                t_amt = float(tuition or 0)
                i_fee = float(id_fee or 0)
                d_fee = float(dmc_fee or 0)
                e_fee = float(exam_fee or 0)
                f_fee = float(fund_fee or 0)
                tot_item = t_amt + i_fee + d_fee + e_fee + f_fee

                paid_amt_val = float(paid_amt or 0)
                status_str = str(status or "").strip().upper()
                is_paid = 1 if (status_str == "PAID" or (tot_item > 0 and paid_amt_val >= tot_item)) else (1 if paid_amt_val > 0 and (tot_item - paid_amt_val) <= 0 else 0)
                paid_date_str = str(paid_date)[:10] if paid_date and str(paid_date).strip() not in ("-", "None", "") else (due_date_str if is_paid else None)

                db.execute(
                    """INSERT INTO payments (student_id, installment_no, due_date, tuition_amount,
                                              id_card_fee, dmc_fee, exam_fee, fund_fee, paid, paid_amount, paid_date)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (target_student_id, inst_no_val, due_date_str, t_amt, i_fee, d_fee, e_fee, f_fee,
                     is_paid, paid_amt_val, paid_date_str)
                )
                counts["payments"] += 1

        db.commit()

    flash(
        f"Import complete: {counts['teachers']} teachers, {counts['courses']} courses, "
        f"{counts['students']} students, and {counts['payments']} payments/installments synced successfully.", "success"
    )
    return redirect(url_for("data_tools"))


# Initialize database on app startup (needed for WSGI servers like Gunicorn on Render)
init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host="0.0.0.0", port=port)
